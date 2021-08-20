import pandas as pd 
from openpyxl import load_workbook
import xlsxwriter

xl = pd.ExcelFile("D:\\WINTER_SEMESTER_2020_21\\CSE3021_SOCIAL_AND_INFORMATION_NETWORKS\\New folder\\archive\\IPL_MATCHES_2018_2020_ANALYZED.xlsx")
df = xl.parse("IPL_MATCHES_2018_2020_ANALYZED")

dict1 = {"Chennai Super Kings" :{},"Delhi Capitals":{},"Kings XI Punjab":{},"Kolkata Knight Riders":{},"Mumbai Indians":{},"Rajasthan Royals":{},"Royal Challengers Bangalore":{},"Sunrisers Hyderabad":{}}
team_to_node={"Chennai Super Kings" :0,"Delhi Capitals":1,"Kings XI Punjab":2,"Kolkata Knight Riders":3,"Mumbai Indians":4,"Rajasthan Royals":5,"Royal Challengers Bangalore":6,"Sunrisers Hyderabad":7}
set1 = {"Chennai Super Kings"}
team_names= []
for i in range(0,179):
	set1.add(df['team1'][i])
team_names = sorted(set1)
# outWorkbook1 = xlsxwriter.Workbook("EDGES_TEAMS_DATA_2018_2020.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","source")
# outSheet1.write("B1","target")
# outSheet1.write("C1","type")
# outSheet1.write("D1","weight")
row=1
dict2 = {"Chennai Super Kings" :{}, "Deccan Chargers":{},"Delhi Capitals":{},"Delhi Daredevils":{},"Gujarat Lions":{},"Kings XI Punjab":{},"Kochi Tuskers Kerala":{},"Kolkata Knight Riders":{},"Mumbai Indians":{},"Pune Warriors":{},"Rajasthan Royals":{},"Rising Pune Supergiant":{},"Royal Challengers Bangalore":{},"Sunrisers Hyderabad":{}}

for i in team_names:
	num_match=0
	wins=0
	for k in range(0,179):
		if(df['team1'][k]==i or df['team2'][k]==i):
			num_match+=1
			if(df['winner'][k]==i):
				wins+=1
	dict2[i]['match']=num_match
	dict2[i]['win']=wins

for i in team_names:
	for j in team_names:
		num_matches=0
		num_wins_team2=0
		for k in range(0,179):
			if((df['team1'][k]==i and df['team2'][k]==j) or (df['team1'][k]==j and df['team2'][k]==i)):
				num_matches+=1
				if(df['winner'][k]==j):
					num_wins_team2+=1

		if num_matches!=0:
			dict1[i][j]=num_wins_team2/num_matches
			# outSheet1.write(row,0,team_to_node[i])
			# outSheet1.write(row,1,team_to_node[j])
			# outSheet1.write(row,2,"Directed")
			# outSheet1.write(row,3,dict1[i][j])
			# row+=1

# outWorkbook = xlsxwriter.Workbook("NODES_TEAMS_DATA_2018_2020.xlsx")

# outSheet = outWorkbook.add_worksheet()

# outSheet.write("A1","id")
# outSheet.write("B1","label")
# for item in range(len(team_names)):
# 	outSheet.write(item+1,0,item)
# 	outSheet.write(item+1,1,team_names[item])
# outWorkbook.close()
# outWorkbook1.close()

#Calculating Page Ranks
page_rank_init={"Chennai Super Kings" :1/8,"Delhi Capitals":1/8,"Kings XI Punjab":1/8,"Kolkata Knight Riders":1/8,"Mumbai Indians":1/8,"Rajasthan Royals":1/8,"Royal Challengers Bangalore":1/8,"Sunrisers Hyderabad":1/8}

page_rank_next={"Chennai Super Kings" :1,"Delhi Capitals":1,"Kings XI Punjab":1,"Kolkata Knight Riders":1,"Mumbai Indians":1,"Rajasthan Royals":1,"Royal Challengers Bangalore":1,"Sunrisers Hyderabad":1}

for k in range(0,5000):
	for i in team_names:
		for j in team_names:
			if i!=j:
				S_j_out=0
				Sum_rank=0
				for l in team_names:
					if l!=j:
					
						S_j_out+=dict1[j][l]
				Sum_rank=page_rank_init[j]*dict1[j][i]
				Sum_rank=Sum_rank/S_j_out
		page_rank_next[i]=Sum_rank*(0.85)+0.15/8
	for i in team_names:
		page_rank_init[i]=page_rank_next[i]
sorted_page = sorted(page_rank_next.items(), key=lambda x: x[1],reverse=True)   
for i in sorted_page:
	print(i)



for i in team_names:
	print(i,"           ",dict2[i]['match'],"            ",dict2[i]['win'],"           ",dict2[i]['win']/dict2[i]['match'])


	

