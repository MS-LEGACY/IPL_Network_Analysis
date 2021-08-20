import pandas as pd 
from openpyxl import load_workbook
import xlsxwriter
import tabulate
import math
import math
xl = pd.ExcelFile("D:\\WINTER_SEMESTER_2020_21\\CSE3021_SOCIAL_AND_INFORMATION_NETWORKS\\New folder\\archive\\IPL_BALL_BY_BALL_2018_2020.xlsx")
df = xl.parse("IPL_BALL_2018_2020")

Chennai_Super_Kings=[]
Delhi_Capitals=[]
Kings_XI_Punjab=[]
Kolkata_Knight_Riders=[]
Mumbai_Indians=[]
Rajasthan_Royals=[]
Royal_Challengers_Bangalore=[]
Sunrisers_Hyderabad=[]
set1=set()
for j in range(0,43089):
	if(df['batting_team'][j]=="Mumbai Indians"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Mumbai_Indians=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Delhi Capitals"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Delhi_Capitals=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Kings XI Punjab"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Kings_XI_Punjab=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Kolkata Knight Riders"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Kolkata_Knight_Riders=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Chennai Super Kings"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Chennai_Super_Kings=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Rajasthan Royals"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Rajasthan_Royals=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Royal Challengers Bangalore"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Royal_Challengers_Bangalore=list(set1)
set1.clear()
for j in range(0,43089):
	if(df['batting_team'][j]=="Sunrisers Hyderabad"):
		set1.add(df['batsman'][j])
		set1.add(df['non_striker'][j])
Sunrisers_Hyderabad=list(set1)
set1.clear()
Chennai_Super_Kings.sort()
Delhi_Capitals.sort()
Kings_XI_Punjab.sort()
Kolkata_Knight_Riders.sort()
Mumbai_Indians.sort()
Rajasthan_Royals.sort()
Royal_Challengers_Bangalore.sort()
Sunrisers_Hyderabad.sort()
# outWorkbook1 = xlsxwriter.Workbook("CSK_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=0
# for i in range(0,len(Chennai_Super_Kings)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Chennai_Super_Kings[i])
# 	row+=1
# 	print(i,Chennai_Super_Kings[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("CSK_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","node1")
# outSheet2.write("B1","node2")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=0
# for i in range(0,len(Chennai_Super_Kings)):
# 	for j in range(0,len(Chennai_Super_Kings)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Chennai_Super_Kings[i]!=Chennai_Super_Kings[j]):
# 				if(df['batting_team'][k]=="Chennai Super Kings"):
# 					if((df['batsman'][k]==Chennai_Super_Kings[i] and df['non_striker'][k]== Chennai_Super_Kings[j]) or (df['batsman'][k]==Chennai_Super_Kings[j] and df['non_striker'][k]==Chennai_Super_Kings[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Chennai_Super_Kings[i],"  ",j,"-",Chennai_Super_Kings[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")

# =============DELHI CAPITALS=======================

# outWorkbook1 = xlsxwriter.Workbook("DC_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=0
# for i in range(0,len(Delhi_Capitals)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Delhi_Capitals[i])
# 	row+=1
# 	print(i,Delhi_Capitals[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("DC_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","node1")
# outSheet2.write("B1","node2")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=1
# for i in range(0,len(Delhi_Capitals)):
# 	for j in range(0,len(Delhi_Capitals)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Delhi_Capitals[i]!=Delhi_Capitals[j]):
# 				if(df['batting_team'][k]=="Delhi Capitals"):
# 					if((df['batsman'][k]==Delhi_Capitals[i] and df['non_striker'][k]== Delhi_Capitals[j]) or (df['batsman'][k]==Delhi_Capitals[j] and df['non_striker'][k]==Delhi_Capitals[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Delhi_Capitals[i],"  ",j,"-",Delhi_Capitals[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")

# ================================KINGS XI PUNJAB===========================================

# outWorkbook1 = xlsxwriter.Workbook("KXIPun_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=1
# for i in range(0,len(Kings_XI_Punjab)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Kings_XI_Punjab[i])
# 	row+=1
# 	print(i,Kings_XI_Punjab[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("KXIPun_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","node1")
# outSheet2.write("B1","node2")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=1
# for i in range(0,len(Kings_XI_Punjab)):
# 	for j in range(0,len(Kings_XI_Punjab)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Kings_XI_Punjab[i]!=Kings_XI_Punjab[j]):
# 				if(df['batting_team'][k]=="Kings XI Punjab"):
# 					if((df['batsman'][k]==Kings_XI_Punjab[i] and df['non_striker'][k]== Kings_XI_Punjab[j]) or (df['batsman'][k]==Kings_XI_Punjab[j] and df['non_striker'][k]==Kings_XI_Punjab[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Kings_XI_Punjab[i],"  ",j,"-",Kings_XI_Punjab[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")


# ================================================Kolkata Knight Riders========================================

# outWorkbook1 = xlsxwriter.Workbook("KKR_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=1
# for i in range(0,len(Kolkata_Knight_Riders)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Kolkata_Knight_Riders[i])
# 	row+=1
# 	print(i,Kolkata_Knight_Riders[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("KKR_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","node1")
# outSheet2.write("B1","node2")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=1
# for i in range(0,len(Kolkata_Knight_Riders)):
# 	for j in range(0,len(Kolkata_Knight_Riders)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Kolkata_Knight_Riders[i]!=Kolkata_Knight_Riders[j]):
# 				if(df['batting_team'][k]=="Kolkata Knight Riders"):
# 					if((df['batsman'][k]==Kolkata_Knight_Riders[i] and df['non_striker'][k]== Kolkata_Knight_Riders[j]) or (df['batsman'][k]==Kolkata_Knight_Riders[j] and df['non_striker'][k]==Kolkata_Knight_Riders[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Kolkata_Knight_Riders[i],"  ",j,"-",Kolkata_Knight_Riders[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")

# ============================================MUMBAI INDIANS===========================================


# outWorkbook1 = xlsxwriter.Workbook("MI_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=1
# for i in range(0,len(Mumbai_Indians)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Mumbai_Indians[i])
# 	row+=1
# 	print(i,Mumbai_Indians[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("MI_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","node1")
# outSheet2.write("B1","node2")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=1
# for i in range(0,len(Mumbai_Indians)):
# 	for j in range(0,len(Mumbai_Indians)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Mumbai_Indians[i]!=Mumbai_Indians[j]):
# 				if(df['batting_team'][k]=="Mumbai Indians"):
# 					if((df['batsman'][k]==Mumbai_Indians[i] and df['non_striker'][k]== Mumbai_Indians[j]) or (df['batsman'][k]==Mumbai_Indians[j] and df['non_striker'][k]==Mumbai_Indians[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Mumbai_Indians[i],"  ",j,"-",Mumbai_Indians[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")


# ====================================================RAJASTHAN ROYALS=============================

# outWorkbook1 = xlsxwriter.Workbook("RR_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=1
# for i in range(0,len(Rajasthan_Royals)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Rajasthan_Royals[i])
# 	row+=1
# 	print(i,Rajasthan_Royals[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("RR_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","node1")
# outSheet2.write("B1","node2")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=1
# for i in range(0,len(Rajasthan_Royals)):
# 	for j in range(0,len(Rajasthan_Royals)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Rajasthan_Royals[i]!=Rajasthan_Royals[j]):
# 				if(df['batting_team'][k]=="Rajasthan Royals"):
# 					if((df['batsman'][k]==Rajasthan_Royals[i] and df['non_striker'][k]== Rajasthan_Royals[j]) or (df['batsman'][k]==Rajasthan_Royals[j] and df['non_striker'][k]==Rajasthan_Royals[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Rajasthan_Royals[i],"  ",j,"-",Rajasthan_Royals[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")


# ===================================================ROYAL CHALLENGERS========================================

# outWorkbook1 = xlsxwriter.Workbook("RCB_Nodes_Partnership.xlsx")

# outSheet1 = outWorkbook1.add_worksheet()
# outSheet1.write("A1","id")
# outSheet1.write("B1","label")
# row=1
# for i in range(0,len(Royal_Challengers_Bangalore)):
# 	outSheet1.write(row,0,i)
# 	outSheet1.write(row,1,Royal_Challengers_Bangalore[i])
# 	row+=1
# 	print(i,Royal_Challengers_Bangalore[i])	
# outWorkbook1.close()
# outWorkbook2 = xlsxwriter.Workbook("RCB_Partnership.xlsx")

# outSheet2 = outWorkbook2.add_worksheet()
# outSheet2.write("A1","source")
# outSheet2.write("B1","target")
# outSheet2.write("C1","type")
# outSheet2.write("D1","weight")
# row=1
# for i in range(0,len(Royal_Challengers_Bangalore)):
# 	for j in range(0,len(Royal_Challengers_Bangalore)):
# 		partnership_scores=0
# 		for k in range(0,43089):
# 			if(Royal_Challengers_Bangalore[i]!=Royal_Challengers_Bangalore[j]):
# 				if(df['batting_team'][k]=="Royal Challengers Bangalore"):
# 					if((df['batsman'][k]==Royal_Challengers_Bangalore[i] and df['non_striker'][k]== Royal_Challengers_Bangalore[j]) or (df['batsman'][k]==Royal_Challengers_Bangalore[j] and df['non_striker'][k]==Royal_Challengers_Bangalore[i])):
# 						partnership_scores+=df['total_runs'][k]
# 		if(partnership_scores > 0):
# 			print(i,"-",Royal_Challengers_Bangalore[i],"  ",j,"-",Royal_Challengers_Bangalore[j]," ",partnership_scores)
# 			outSheet2.write(row,0,i)
# 			outSheet2.write(row,1,j)
# 			outSheet2.write(row,2,"Undirected")
# 			outSheet2.write(row,3,partnership_scores)
# 			row+=1
# outWorkbook2.close()		
# print("Hello World")



# ===========================================================SUNRISERS HYDERABAD=======================================

outWorkbook1 = xlsxwriter.Workbook("SRH_Nodes_Partnership.xlsx")

outSheet1 = outWorkbook1.add_worksheet()
outSheet1.write("A1","id")
outSheet1.write("B1","label")
row=1
for i in range(0,len(Sunrisers_Hyderabad)):
	outSheet1.write(row,0,i)
	outSheet1.write(row,1,Sunrisers_Hyderabad[i])
	row+=1
	print(i,Sunrisers_Hyderabad[i])	
outWorkbook1.close()
outWorkbook2 = xlsxwriter.Workbook("SRH_Partnership.xlsx")

outSheet2 = outWorkbook2.add_worksheet()
outSheet2.write("A1","source")
outSheet2.write("B1","target")
outSheet2.write("C1","type")
outSheet2.write("D1","weight")
row=1
for i in range(0,len(Sunrisers_Hyderabad)):
	for j in range(0,len(Sunrisers_Hyderabad)):
		partnership_scores=0
		for k in range(0,43089):
			if(Sunrisers_Hyderabad[i]!=Sunrisers_Hyderabad[j]):
				if(df['batting_team'][k]=="Sunrisers Hyderabad"):
					if((df['batsman'][k]==Sunrisers_Hyderabad[i] and df['non_striker'][k]== Sunrisers_Hyderabad[j]) or (df['batsman'][k]==Sunrisers_Hyderabad[j] and df['non_striker'][k]==Sunrisers_Hyderabad[i])):
						partnership_scores+=df['total_runs'][k]
		if(partnership_scores > 0):
			print(i,"-",Sunrisers_Hyderabad[i],"  ",j,"-",Sunrisers_Hyderabad[j]," ",partnership_scores)
			outSheet2.write(row,0,i)
			outSheet2.write(row,1,j)
			outSheet2.write(row,2,"Undirected")
			outSheet2.write(row,3,partnership_scores)
			row+=1
outWorkbook2.close()		
print("Hello World")